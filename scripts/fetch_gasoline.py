"""
石油製品小売市況調査（週次 xlsx）をダウンロード・パースし、
data/gasoline_prices.json（全国平均・週次推移）を更新するスクリプト。

石油版 (scripts/fetch_pdf.py) / LPG版 (scripts/fetch_lpg.py) の姉妹スクリプト。
相違点:

    1. 石油製品価格調査は資源エネルギー庁 pl007 で「毎週」公表され、
       給油所小売価格（店頭現金・消費税込み）を Excel(.xlsx) で提供する。
       ファイルは /pl007/xlsx/<YYMMDD>.xlsx（例: 260729.xlsx）。
       サフィックス付き（...s5 / ...k / ...a / ...o 等）は軽油インタンクや
       産業用など別調査なので採らない。無サフィックスの週次 SS 小売価格のみ。
    2. 保存する値は「全国平均（参考値）」のみ。都道府県別・地域別は採らない
       （フロントは全国の週次推移グラフとして表示する方針）。
    3. 抽出対象油種は per-liter の道路燃料 3 種:
           ハイオク(highOctane) / レギュラー(regular) / 軽油(diesel)
       ※ 灯油は「円/18L」と「円/L」が混在し全国列の位置も異なるため v1 では対象外。

抽出する値（各週の1行）:
    { "asOf": "2026-07-27", "regular": 170.1, "highOctane": 180.9, "diesel": 159.4 }

    asOf は各油種ブロックの「調査日」。price は「全国（円／ﾘｯﾄﾙ）（参考値）」列。

出口コード:
    0 — 成功（変更あり / 変更なしの両方とも0）
    1 — 失敗（ダウンロード/パース/検証エラー）

使い方:
    python scripts/fetch_gasoline.py [--dry-run] [--xlsx-path path/to/file.xlsx]

注意（WAF）:
    経産省サイトは AWS WAF の JS チャレンジを返すことがある。curl-cffi の
    impersonate で実ブラウザ fingerprint を模倣し、results.html を先に踏んで
    セッションを確立してから xlsx を取得する。それでも弾かれる場合はリトライ
    /backoff で待つ（fetch_pdf / fetch_lpg と同じ運用）。
"""

from __future__ import annotations

import argparse
import datetime
import io
import re
import sys
import time
from pathlib import Path

import openpyxl
from curl_cffi import requests as crequests
from lib.io import read_json, write_json
from lib.paths import DATA_DIR

GASOLINE_PATH = DATA_DIR / "gasoline_prices.json"

PL007_BASE = "https://www.enecho.meti.go.jp/statistics/petroleum_and_lpgas/pl007/"
RESULTS_URL = PL007_BASE + "results.html"
SITE_ORIGIN = "https://www.enecho.meti.go.jp"

TMP_XLSX_PATH = DATA_DIR / ".gasoline_weekly.xlsx"

# 経産省サイトの WAF は TLS フィンガープリント (JA3) / HTTP/2 SETTINGS まで見る。
IMPERSONATE_BROWSER = "chrome"

# 経済産業局別シート（全国参考値が載る）。都道府県別シートは採らない。
REGION_SHEET = "公表資料（経済産業局別）"

# results.html 内の週次 SS 小売価格 xlsx。無サフィックス（YYMMDD のみ）に限定する。
XLSX_LINK_RE = re.compile(
    r'href="(?P<href>[^"]*?/pl007/xlsx/(?P<date>\d{6})\.xlsx)"',
    re.IGNORECASE,
)

# シート上の油種ラベル → JSON キー。全角スペース除去後に完全一致で判定。
PRODUCTS = {
    "ハイオク": "highOctane",
    "レギュラー": "regular",
    "軽油": "diesel",
}

# per-liter の妥当域（円/L）。これを外れたら異常値として遮断する。
PRICE_MIN = 100.0
PRICE_MAX = 300.0


def _norm(v: object) -> str:
    """セル値を全角スペース除去した文字列に正規化。"""
    return str(v).replace("　", "").replace(" ", "")


def _date_from_filename(token: str) -> datetime.date:
    """xlsx ファイル名の日付トークン(YYMMDD)を date に変換。"""
    y, m, d = 2000 + int(token[0:2]), int(token[2:4]), int(token[4:6])
    return datetime.date(y, m, d)


def _iso(v: object) -> str:
    """datetime / date セルを ISO 日付文字列に。"""
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    raise TypeError(f"not a date cell: {v!r}")


def resolve_latest_xlsx_url(
    session: crequests.Session, *, timeout: int = 60, retries: int = 3, backoff: float = 5.0
) -> tuple[str, datetime.date]:
    """results.html を取得し、最新（ファイル名日付が最大）の週次 xlsx URL を返す。"""
    last_err: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            resp = session.get(RESULTS_URL, timeout=timeout)
            resp.raise_for_status()
            html = resp.text
            candidates: list[tuple[datetime.date, str]] = []
            for m in XLSX_LINK_RE.finditer(html):
                href = m.group("href")
                pub = _date_from_filename(m.group("date"))
                url = href if href.startswith("http") else SITE_ORIGIN + href
                candidates.append((pub, url))
            if not candidates:
                raise RuntimeError("no weekly xlsx link found in results.html")
            pub, url = max(candidates, key=lambda t: t[0])
            return url, pub
        except Exception as e:  # noqa: BLE001 - リトライして最終的に abort
            last_err = e
            print(
                f"[gas] resolve attempt {attempt}/{retries} failed: {type(e).__name__}: {e}",
                file=sys.stderr,
            )
            if attempt < retries:
                time.sleep(backoff * attempt)
    raise RuntimeError(f"all {retries} resolve attempts failed") from last_err


def download_xlsx(
    session: crequests.Session,
    url: str,
    dest: Path,
    *,
    timeout: int = 60,
    retries: int = 4,
    backoff: float = 5.0,
) -> None:
    """xlsx をダウンロード。WAF の HTML チャレンジ（先頭が PK でない）はリトライ。"""
    dest.parent.mkdir(parents=True, exist_ok=True)
    last_err: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            resp = session.get(url, headers={"Referer": RESULTS_URL}, timeout=timeout)
            resp.raise_for_status()
            data = resp.content
            if data[:2] != b"PK":  # xlsx は ZIP(PK)。HTML チャレンジは弾く。
                raise RuntimeError("response is not an xlsx (WAF challenge?)")
            dest.write_bytes(data)
            return
        except Exception as e:  # noqa: BLE001
            last_err = e
            print(
                f"[gas] download attempt {attempt}/{retries} failed: {type(e).__name__}: {e}",
                file=sys.stderr,
            )
            if attempt < retries:
                time.sleep(backoff * attempt)
    raise RuntimeError(f"all {retries} download attempts failed") from last_err


def load_grid(xlsx_path: Path) -> list[tuple]:
    """経済産業局別シートを行タプルのリストとして読み込む。"""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_path.read_bytes()), data_only=True)
    if REGION_SHEET not in wb.sheetnames:
        raise RuntimeError(
            f"sheet {REGION_SHEET!r} not found; got {wb.sheetnames}"
        )
    ws = wb[REGION_SHEET]
    return list(ws.iter_rows(values_only=True))


def find_national_col(grid: list[tuple]) -> int:
    """ヘッダ行から「全国」を含む列（＝参考値列）のインデックスを返す。"""
    for row in grid[:12]:
        for ci, v in enumerate(row):
            if v is not None and "全国" in _norm(v):
                return ci
    raise RuntimeError("national ('全国') column not found in header")


def parse_grid(grid: list[tuple]) -> list[dict]:
    """
    経済産業局別シートを週次レコードに変換。

    各油種ブロックの「調査日」行から (asOf, 油種, 全国参考値) を拾い、
    asOf ごとに { asOf, regular, highOctane, diesel } へピボットする。
    """
    ncol = find_national_col(grid)
    current: str | None = None
    by_date: dict[str, dict] = {}

    for row in grid:
        # 油種ラベルの検出（ブロック先頭行にだけ現れる）。
        for v in row:
            if v is None:
                continue
            key = _norm(v)
            if key in PRODUCTS:
                current = PRODUCTS[key]
                break

        if current is None:
            continue

        # 調査日（date セル）を含む行だけを価格行として扱う。
        as_of: str | None = None
        for v in row:
            if isinstance(v, (datetime.datetime, datetime.date)) and not isinstance(
                v, bool
            ):
                as_of = _iso(v)
                break
        if as_of is None:
            continue

        price = row[ncol] if ncol < len(row) else None
        if not isinstance(price, (int, float)) or isinstance(price, bool):
            continue
        if price <= 0:
            continue

        rec = by_date.setdefault(as_of, {"asOf": as_of})
        rec[current] = round(float(price), 1)

    rows = [by_date[k] for k in sorted(by_date)]
    if not rows:
        raise RuntimeError("no price rows parsed; xlsx layout may have changed")
    return rows


def validate(rows: list[dict]) -> None:
    """各週レコードの妥当性（日付・値域・油種の存在）を検証。"""
    wanted = set(PRODUCTS.values())
    for r in rows:
        try:
            datetime.date.fromisoformat(r["asOf"])
        except (ValueError, KeyError) as e:
            raise RuntimeError(f"invalid asOf: {r}") from e
        present = wanted & r.keys()
        if not present:
            raise RuntimeError(f"row has no product prices: {r}")
        for key in present:
            p = r[key]
            if not (PRICE_MIN <= p <= PRICE_MAX):
                raise RuntimeError(f"{key} out of range on {r['asOf']}: {p}")


def load_existing() -> list[dict]:
    if not GASOLINE_PATH.exists():
        return []
    return read_json(GASOLINE_PATH)


def merge(existing: list[dict], new: list[dict]) -> tuple[list[dict], int, int]:
    """asOf キーで統合。新規 xlsx の値を正典として上書き。"""
    by_date: dict[str, dict] = {r["asOf"]: r for r in existing}
    added = updated = 0
    for r in new:
        key = r["asOf"]
        if key not in by_date:
            by_date[key] = r
            added += 1
        elif by_date[key] != r:
            by_date[key] = r
            updated += 1
    merged = sorted(by_date.values(), key=lambda x: x["asOf"])
    return merged, added, updated


def _make_session() -> crequests.Session:
    return crequests.Session(impersonate=IMPERSONATE_BROWSER)


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--keep-xlsx", action="store_true")
    parser.add_argument(
        "--xlsx-path",
        type=Path,
        default=None,
        help="ローカル xlsx を使う（テスト用。指定時はダウンロードしない）",
    )
    args = parser.parse_args(argv)

    if args.xlsx_path is None:
        session = _make_session()
        try:
            url, pub_date = resolve_latest_xlsx_url(session)
            print(f"[gas] latest xlsx: {url} (published {pub_date})")
            # results.html を踏んでからセッションで取得（WAF セッション確立）。
            download_xlsx(session, url, TMP_XLSX_PATH)
        except Exception as e:  # noqa: BLE001
            print(f"[gas] fetch failed: {type(e).__name__}: {e}", file=sys.stderr)
            return 1
        xlsx_path = TMP_XLSX_PATH
    else:
        xlsx_path = args.xlsx_path
        print(f"[gas] using local xlsx: {xlsx_path}")
        if not xlsx_path.exists():
            print(f"[gas] file not found: {xlsx_path}", file=sys.stderr)
            return 1

    try:
        grid = load_grid(xlsx_path)
        rows = parse_grid(grid)
        validate(rows)
    except Exception as e:  # noqa: BLE001
        print(f"[gas] parse/validate failed: {type(e).__name__}: {e}", file=sys.stderr)
        return 1
    finally:
        if not args.keep_xlsx and args.xlsx_path is None:
            TMP_XLSX_PATH.unlink(missing_ok=True)

    print(f"[gas] parsed {len(rows)} weekly rows; latest: {rows[-1]}")

    existing = load_existing()
    merged, added, updated = merge(existing, rows)
    print(f"[gas] existing: {len(existing)}, added: {added}, updated: {updated}")

    if added == 0 and updated == 0:
        print("[gas] no changes")
        return 0

    if args.dry_run:
        print("[gas] dry-run; not writing gasoline_prices.json")
        return 0

    write_json(GASOLINE_PATH, merged)
    print(f"[gas] wrote {GASOLINE_PATH} with {len(merged)} rows")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
